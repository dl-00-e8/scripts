#!python3
# -*- coding:utf-8 -*-

import openpyxl
import datetime

except_lists = [

]

class cell_xlsx:
    def __init__(self, descript, ip_lists):
        self.description = descript
        self.input_lists = ip_lists
        self.now = datetime.datetime.now()
        self.now_time = str(self.now.year) + "-" + str(self.now.month) + "-" + str(self.now.day) + "-" + str(self.now.hour) + "��" + str(self.now.minute) + "��"
        self.file_name = self.now_time + "-blacklist.xlsx"
    

    def check(self):
        self.ip_list = []
        for ip in self.input_lists:
            if ip in except_lists:
                continue
            self.ip_list.append(ip)
    

    def write(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

        self.sheet['A1'] = '����� �ּ�'
        self.sheet['B1'] = '����� ��Ʈ'
        self.sheet['C1'] = '������ �ּ�'
        self.sheet['D1'] = '������ ��Ʈ'
        self.sheet['E1'] = '��������'
        self.sheet['F1'] = '����'

        for i in range(2, len(self.ip_list) + 1):
            self.sheet.cell(column = 1, row = i, value = 'ANY')
            self.sheet.cell(column = 2, row = i, value = 'ANY')
            self.sheet.cell(column = 3, row = i, value = self.ip_list[i - 2])
            self.sheet.cell(column = 4, row = i, value = 'ANY')
            self.sheet.cell(column = 5, row = i, value = 'ANY')
            self.sheet.cell(column = 6, row = i, value = self.description)
        
        self.wb.save(self.file_name)
        self.wb.close()